import time, traceback, requests, json
from utils import clean_text

def count_kws_in_org(org:str, key_words:list, sum_threshold = 2):
    # org = clean_text(org)
    try: 
        data = {
            "from":0,
            "size":10000,
            "sort": { "_id": { "order": "asc" }}
        }
        headers = {'Content-Type': 'application/json'}
        r = requests.get(f"http://localhost:9200/{org.lower()}/_search", headers = headers, data = json.dumps(data))

        lines = []
        _j = json.loads(r.text)['hits']['hits']
        for item in _j:

            index = item['_id']
            name = item['_source']['Name']
            if name.startswith('Proceedings of'): continue
            org = item['_source']['Organization']
            year = item['_source']['Year']
            text = item['_source']['Text']

            one_line = [index, name, org, year]
            count_res = [text.count(i) for i in key_words]
            if sum(count_res) <= sum_threshold: continue
            one_line.extend(count_res)

            lines.append(one_line)
        return lines
    except:
        traceback.print_exc()
        return f'ERROR: {org} {key_words}'

from utils import org_years_dict
def count_to_excel(orgs = org_years_dict.keys(), key_words = None, split_to_sheet = True):
    
    if not key_words: return
    start_time = time.time()
    key_words = [clean_text(i) for i in key_words]
    res_dict = {}
    for org in orgs:
        res_dict[org] = count_kws_in_org(org, key_words)
        
    from openpyxl import Workbook
    wb = Workbook()
    _first_line = ['Index', 'Name', 'Org', 'Year']
    _first_line.extend(key_words)
    if split_to_sheet:
        for _org in orgs:
            sheet = wb.create_sheet(_org, index=0)
            sheet.append(_first_line)
            for line in res_dict[_org]:
                sheet.append(line)
    
    # write to single sheet.  
    else:    
        sheet = wb.create_sheet('res', index=0)
        sheet.append(_first_line)
        for _org in orgs:
            for line in res_dict[_org]:
                sheet.append(line)
    
    
    _time_now = time.strftime("%Y-%m-%d_%H-%M-%S", time.localtime())
    _orgs = '_'.join(orgs)
    wb.save(f'RES_{_orgs}_{_time_now}.xlsx')
    end_time = time.time()
    print(f'DONE! time cost: {end_time - start_time:.2f}s')


if __name__ == "__main__":
    # count_to_excel(['ACL', 'EMNLP'], ['webquestion', 'GCN'])
    # kws = ['Free917', 'WebQuestions', 'WebquestionSP', 'ComplexQuestions', 'GraphQuestions',
    #         'LC-QuAD', 'QALD', 'QALD-7', 'Factoid', 'KGQA', 'KBQA', 'λ-DCS', 'GCN']
    # kws = ['multi-hop', 'multi-hop reasoning', 'Reading Comprehension', 'Question Answering', 'Machine Reading',
    #          'KGQA', 'KBQA', 'GCN']

    kws = ['Graph Neural Network', 'GCN', 'knowledge graph', 'WebQuestionsSP', 'WebQSP', 'WebQuestions', 'COMPLEXWEBQUESTIONS', 'WikiMovies', 'LC-QuAD', 'QALD', 'GeoQuery', \
        'GEO ', 'SPIDER ', 'ATIS ', 'OVERNIGHT ']

    # kws = ['relation extraction', 'relation classification','graph neural network', 'graph convolutional network', 'TACRED', 'SemEval-2010 Task 8', 'new york Times']

    # kws = ['Graph Neural Network', 'GCN', 'semantic role labeling', 'CoNLL 2005', 'CoNLL 2009', 'CoNLL 2012']

    count_to_excel(key_words = kws, split_to_sheet=False)







